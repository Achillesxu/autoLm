# -*- coding: utf-8 -*-
"""
@Time : 2024 9月 02 13:49
@Author : xushiyin
@Mobile : 18682193124
@desc :
"""
import io
import json

from rich import print  # noqa
from rich.traceback import Traceback  # noqa
from openpyxl import load_workbook
from handlers.utils import get_cur_user

class ExcelHandler:
    def __init__(self, f_path: str | io.BytesIO, sheet_name: str | None = None):
        self.cur_user = get_cur_user()
        self.f_path = f'C:/Users/{self.cur_user}/Desktop/{f_path}'
        self.sheet_name = sheet_name
        self.wb = None
        self.ws = None


    def load_excel(self):
        self.wb = load_workbook(self.f_path, data_only=True, read_only=True)

    def get_all_sheet_names(self):
        return self.wb.sheetnames

    @staticmethod
    def find_valid_key_value(in_dict, in_key):
        """only ascii letter"""
        if in_key in in_dict:
            return in_dict[in_key]
        else:
            ss = in_key
            while True:
                ss = chr(ord(ss) - 1)
                if ss in in_dict:
                    return in_dict[ss]

    def get_cur_sheet(self):
        print(f'读取excel: [red]{self.f_path}[/red]')
        self.load_excel()
        if self.sheet_name is None:
            s_names = self.get_all_sheet_names()
            self.sheet_name = s_names[0]
            print(f'默认读取 sheet: [red]{self.sheet_name}[/red]')
        else:
            print(f'读取 sheet: [red]{self.sheet_name}[/red]')

        cur_sheet = self.wb[self.sheet_name]
        return cur_sheet

    def read_psd_path(self):
        cur_sheet = self.get_cur_sheet()
        row_header = list()
        row_layers = list()
        row_data = list()

        try:
            for row in cur_sheet.iter_rows(min_row=2, max_row=100, min_col=1, max_col=2, values_only=True):
                if row:
                    row_data.append(list(row))
        except Exception: # noqa
            print(Traceback())
            return
        file_path = list()

        for row in row_data:
            f_dir = row[0].replace('\\', '/')
            file_path.append(f'{f_dir}/{row[1]}.psd')

        for cell in cur_sheet[1]:
            if cell.value:
                row_header.append(cell.value)

        for row in row_data:
            row_layers.append([row_header[idx] for idx, d in enumerate(row[2:], start=2) if d])

        return file_path, row_layers

    def read_sheet_data(self):
        cur_sheet = self.get_cur_sheet()
        row_header = list()
        row_data = list()

        for cell in cur_sheet[1]:
            if cell.value:
                row_header.append(cell.value)

        try:
            for row in cur_sheet.iter_rows(min_row=2, max_row=100, min_col=1, max_col=len(row_header), values_only=True):
                if row:
                    row_data.append(list(row))
        except Exception: # noqa
            print(Traceback())
            return

        data_dict = {
            'psds': {},
            "output": {
                "dir": f"C:/Users/{self.cur_user}/Desktop/jpgs",
                "quality": 8
            }
        }

        for row in row_data:
            tmp = dict()
            for idx, col in enumerate(row[2:], start=2):
                if col:
                    tmp[row_header[idx]] = col
            file_path = row[0].replace('\\', '/')
            data_dict['psds'][f'{file_path}/{row[1]}.psd'] = tmp

        with open(f'C:/code/autoLm/TextLayerData.json', "w", encoding='utf8') as f:
            json.dump(data_dict, f, ensure_ascii=False, indent=4)

        print(f'桌面已经生成 [red]C:/code/autoLm/TextLayerData.json[/red] 文件')


if __name__ == '__main__':
    eh = ExcelHandler('修改.xlsx')
    eh.read_sheet_data()
